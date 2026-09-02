#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""信息脱敏上云 SOP —— 识别规则与常量模块（desen_rules）。

本模块为纯数据/纯函数（正则、白名单、校验器、掩码规则表），无运行时状态、
无对外依赖，供 desensitize.py 导入复用。**不可单独运行**。
"""

import re

# ---------------------------------------------------------------------------
# 1. 敏感字段正则（中文优先）
# ---------------------------------------------------------------------------
# 说明：身份证 / 手机号先处理，其数字会被替换为 '*'，后续数字类正则不会重复命中。
#
# 边界设计：原先用 \b 词边界，但 Python 正则里 CJK 也算"词字符"，
# 导致"数字紧挨中文"（如 手机138...、身份证110...）时 机1/证1 之间无词边界、
# 标识符漏检。现改用自定义边界：
#   _BOUND_L = 左侧不得是 ASCII 字母/数字（CJK、标点、空白、串首 均视为边界）
#   _BOUND_R = 右侧不得是 ASCII 字母/数字（CJK、标点、空白、串尾 均视为边界）
# 效果：① 数字紧邻中文可被正确识别（召回提升）；
#       ② 仍阻止"在更长字母/数字串内部"误匹配（如 a138... / 138…9 不误命中）。
_BOUND_L = r"(?<![0-9A-Za-z])"
_BOUND_R = r"(?![0-9A-Za-z])"
# 跨境标识（iban/swift/vat/intl_phone）用更严边界：额外排除下划线，避免把
# 变量名里的全大写英文词误判（如 WEBHOOK_CALLBACK_IP 的 CALLBACK 被当 SWIFT）。
_BOUND_L_ = r"(?<![0-9A-Za-z_])"
_BOUND_R_ = r"(?![0-9A-Za-z_])"

PATTERNS = {
    "id_card": re.compile(_BOUND_L + r"[1-9]\d{5}(?:18|19|20)\d{2}(?:0[1-9]|1[0-2])(?:0[1-9]|[12]\d|3[01])\d{3}[\dXx]" + _BOUND_R),
    "phone": re.compile(_BOUND_L + r"1[3-9]\d{9}" + _BOUND_R),
    "bank_card": re.compile(_BOUND_L + r"\d{16,19}" + _BOUND_R),
    "ip": re.compile(_BOUND_L + r"(?:(?:25[0-5]|2[0-4]\d|1?\d?\d)\.){3}(?:25[0-5]|2[0-4]\d|1?\d?\d)" + _BOUND_R),
    "email": re.compile(_BOUND_L + r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}" + _BOUND_R),
    "plate": re.compile(_BOUND_L + r"[京津沪渝冀豫云辽黑湘皖鲁新苏浙赣鄂桂甘晋蒙陕吉闽贵粤青藏川宁琼使领][A-Z][A-Z0-9]{5,6}" + _BOUND_R),
    "passport": re.compile(_BOUND_L + r"[EGDSH]\d{8}" + _BOUND_R + r"|" + _BOUND_L + r"[a-zA-Z]\d{9}" + _BOUND_R),
    # JWT（JSON Web Token）：eyJ 开头、三段 base64url。日志/配置中高频明文泄漏点，
    # 特征极强（eyJ 前缀 + 三段点分隔），误报风险低，对所有文本生效。
    "jwt": re.compile(_BOUND_L + r"eyJ[A-Za-z0-9_-]{8,}\.[A-Za-z0-9_-]{8,}\.[A-Za-z0-9_-]{8,}" + _BOUND_R),
    # 跨境银行/税务标识（匹配后须经校验：IBAN 用 MOD-97，SWIFT/VAT 用 ISO 国家码，见下方校验函数）。
    # IBAN：2 字母国家码 + 2 位校验码 + 11~30 位字母数字（总长 15~34）。
    "iban": re.compile(_BOUND_L_ + r"[A-Z]{2}\d{2}[A-Z0-9]{11,30}" + _BOUND_R_),
    # SWIFT/BIC：4 字母银行码 + 2 字母国家码 + 2 字母数字地区码 + 可选 3 位分行码（8 或 11 位）。
    "swift": re.compile(_BOUND_L_ + r"[A-Z]{4}[A-Z]{2}[A-Z0-9]{2}(?:[A-Z0-9]{3})?" + _BOUND_R_),
    # VAT 增值税号：2 字母国家码 + 8~12 位数字（各国格式不一，靠欧盟/EEA 国家码校验降误报）。
    "vat": re.compile(_BOUND_L_ + r"[A-Z]{2}\d{8,12}" + _BOUND_R_),
    # 国际电话：+ 号 + 国家码(1~3 位) + 数字组（空格/横线/点分隔，含括号）。特征强、误报低。
    "intl_phone": re.compile(_BOUND_L_ + r"\+[1-9]\d{0,3}[\d\s.\-()]{4,18}\d" + _BOUND_R_),
}

# ISO 3166-1 alpha-2 国家码（249 个，用于 SWIFT/VAT 的国别校验，降低误报）
_ISO_ALPHA2 = set("""AD AE AF AG AI AL AM AO AQ AR AS AT AU AW AX AZ BA BB BD BE BF BG BH BI BJ
BL BM BN BO BQ BR BS BT BV BW BY BZ CA CC CD CF CG CH CI CK CL CM CN CO CR CU CV CW CX CY CZ
DE DJ DK DM DO DZ EC EE EG EH ER ES ET FI FJ FK FM FO FR GA GB GD GE GF GG GH GI GL GM GN GP
GQ GR GS GT GU GW GY HK HM HN HR HT HU ID IE IL IM IN IO IQ IR IS IT JE JM JO JP KE KG KH KI
KM KN KP KR KW KY KZ LA LB LC LI LK LR LS LT LU LV LY MA MC MD ME MF MG MH MK ML MM MN MO MP
MQ MR MS MT MU MV MW MX MY MZ NA NC NE NF NG NI NL NO NP NR NU NZ OM PA PE PF PG PH PK PL PM
PN PR PS PT PW PY QA RE RO RS RU RW SA SB SC SD SE SG SH SI SJ SK SL SM SN SO SR SS ST SV SX
SY SZ TC TD TF TG TH TJ TK TL TM TN TO TR TT TV TW TZ UA UG UM US UY UZ VA VC VE VG VI VN VU
WF WS YE YT ZA ZM ZW""".split())


def _valid_iban(s: str) -> bool:
    """IBAN MOD-97 校验：前 4 位移到末尾、字母转数字(A=10..Z=35)、mod 97 应 == 1。"""
    if not (15 <= len(s) <= 34):
        return False
    if not re.fullmatch(r"[A-Z]{2}\d{2}[A-Z0-9]+", s):
        return False
    rearranged = s[4:] + s[:4]
    digits = "".join(str(int(ch, 36)) for ch in rearranged)
    try:
        return int(digits) % 97 == 1
    except ValueError:
        return False


def _valid_country(code: str) -> bool:
    """SWIFT/BIC 的国家码是否为合法 ISO 3166-1 alpha-2（SWIFT 覆盖全球）。"""
    return code in _ISO_ALPHA2


# VAT 增值税号的国家码限定为「欧盟 + 欧洲经济区 + 瑞士」（VAT 为欧洲税制），
# 避免把东南亚/北美等「国家码前缀 + 数字」的订单号/流水号（如 SG123456789、
# MY987654321、TH555444333）误判为 VAT。
_EU_VAT_CODES = set(
    "AT BE BG CY CZ DE DK EE ES FI FR GB GR HR HU IE IT LT LU LV MT NL NO PL PT RO "
    "SE SI SK CH IS LI".split()
)


# 各类别的校验器（None 表示无额外校验）；校验失败则不脱敏（保留原样，降低误报）
_KIND_VALIDATOR = {
    "iban": _valid_iban,
    "swift": lambda s: _valid_country(s[4:6]),
    "vat": lambda s: s[0:2] in _EU_VAT_CODES,
}

# 全角 → 半角归一表：数字０-９、大写Ａ-Ｚ、小写ａ-ｚ、全角空格。用于把全角手机/身份证/
# 银行卡等归一为半角后再识别（用户从全角输入法/复制网页常带入全角数字，导致漏检）。
_FULLWIDTH_TABLE = {c: chr(c - 0xFF10 + ord("0")) for c in range(0xFF10, 0xFF1A)}
_FULLWIDTH_TABLE.update({c: chr(c - 0xFF21 + ord("A")) for c in range(0xFF21, 0xFF3B)})
_FULLWIDTH_TABLE.update({c: chr(c - 0xFF41 + ord("a")) for c in range(0xFF41, 0xFF5B)})
_FULLWIDTH_TABLE[0x3000] = " "


def _normalize_fullwidth(text: str) -> str:
    """全角数字/字母/空格归一为半角（仅作用于字符，不改换行等结构）。"""
    return text.translate(_FULLWIDTH_TABLE)


# ---------------------------------------------------------------------------
# 1.1 结构化列解析（--tabular-names）：表格列头英文姓名识别
# ---------------------------------------------------------------------------
# 跨境电商订单导出（CSV/TSV）首行为列头，姓名列的值是英文姓名（Title Case 二词）。
# 纯正则无法区分「John Smith」与「Wireless Earbuds」，但结合列头标签即可可靠识别：
# 仅当列头命中「姓名类标签」且排除「电话/地址/邮箱」等后缀时，才对该列值做英文姓名识别。
# 这是零依赖、纯本地、误报可控的方案（NER 需下载模型，暂不引入，见能力边界）。

# Title Case 英文姓名（含拉丁重音）：首字母大写 + 小写，2~3 词（John Smith / Hans Müller）
_EN_NAME_RE = re.compile(
    r"^[A-ZÀ-ÖØ-Þ][a-zà-öø-ÿ'\-]{1,15}(?:\s+[A-ZÀ-ÖØ-Þ][a-zà-öø-ÿ'\-]{1,15}){1,2}$")

# 姓名列标签（精确匹配；含 name 词根的列头也视为姓名列）
_TABULAR_NAME_LABELS = {
    "name", "fullname", "buyername", "customername", "contactname", "recipientname",
    "consigneename", "shippername", "ownername", "username", "customer", "contact",
    "influencer", "consignee", "shipper", "recipient", "owner", "buyer", "client",
    "买家", "姓名", "联系人", "收件人", "寄件人", "客户", "用户名",
}
# 排除后缀（含这些词的列头不是姓名列，避免 BuyerPhone/ShippingAddress 误判）
_TABULAR_NAME_EXCLUDE = ("phone", "address", "email", "tel", "mobile", "ip", "zip",
                         "city", "country", "电话", "地址", "邮箱", "账号")


def _is_tabular_name_col(header: str) -> bool:
    """判断 CSV/TSV 列头是否为「英文姓名列」（精确匹配标签 + 排除电话/地址/邮箱等后缀）。"""
    h = header.strip().lower().replace("_", " ").replace("-", " ").replace("  ", " ").strip()
    if not h:
        return False
    if any(k in h for k in _TABULAR_NAME_EXCLUDE):
        return False
    if h in _TABULAR_NAME_LABELS or h.endswith("name"):
        return True
    return "姓名" in h or "买家" in h


def _collect_tabular_names(text: str, delimiter: str) -> set:
    """解析 CSV/TSV 文本，收集「姓名列」里的英文姓名（Title Case 二词）。

    返回英文姓名集合（复用 names 机制精确匹配脱敏）。仅收集匹配 _EN_NAME_RE 的值，
    泰文/中文等非拉丁姓名不收集（需 --names 名单兜底）。
    """
    names = set()
    try:
        import csv as _csv
        from io import StringIO
        rows = list(_csv.reader(StringIO(text), delimiter=delimiter))
    except Exception:
        return names
    if not rows:
        return names
    name_cols = [i for i, h in enumerate(rows[0]) if _is_tabular_name_col(h)]
    if not name_cols:
        return names
    for row in rows[1:]:
        for i in name_cols:
            if i < len(row):
                v = row[i].strip()
                if v and _EN_NAME_RE.match(v):
                    names.add(v)
    return names


def _collect_tabular_names_xlsx(path: str) -> set:
    """对 xlsx 文件，用 openpyxl 读各 sheet 首行列头，收集「姓名列」的英文姓名。

    与 _collect_tabular_names（CSV/TSV）对应，补齐 --tabular-names 对 xlsx 的覆盖：
    xlsx 是二进制，抽取为空格分隔文本后列对齐不可靠，故直接读单元格结构做列头解析。
    复用 names 机制精确匹配脱敏；泰文/中文等非拉丁姓名不收集（需 --names）。
    """
    names = set()
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return names
    try:
        for ws in wb.worksheets:
            header = []
            for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                header = [str(c) if c is not None else "" for c in row]
                break
            name_cols = [i for i, h in enumerate(header) if _is_tabular_name_col(h)]
            if not name_cols:
                continue
            for row in ws.iter_rows(min_row=2, values_only=True):
                for i in name_cols:
                    if i < len(row) and row[i] is not None:
                        v = str(row[i]).strip()
                        if v and _EN_NAME_RE.match(v):
                            names.add(v)
    finally:
        try:
            wb.close()
        except Exception:
            pass
    return names


# ---------------------------------------------------------------------------
# 1.1 中文识别增强正则（--cn-enhance 时启用；纯本地，无需联网/模型）
# ---------------------------------------------------------------------------
# 常见复姓（须置于单姓之前优先匹配）
_CN_COMPOUND = ["欧阳", "司马", "上官", "诸葛", "东方", "皇甫", "令狐", "夏侯",
                "宇文", "慕容", "司徒", "端木", "拓跋", "轩辕", "公孙", "长孙",
                "鲜于", "耶律", "完颜", "万俟", "司空", "子车", "颛孙"]
# 常见单姓（百家姓 + 常见姓）
_CN_SINGLE = ("赵钱孙李周吴郑王冯陈褚卫蒋沈韩杨朱秦尤许何吕施张孔曹严华金魏陶姜"
              "戚谢邹喻柏水窦章云苏潘葛奚范彭郎鲁韦昌马苗凤花方俞任袁柳酆鲍史唐"
              "费廉岑薛雷贺倪汤滕殷罗毕郝邬安常乐于时傅皮卞齐康伍余元卜顾孟平黄"
              "和穆萧尹姚邵湛汪祁毛禹狄米贝明臧计伏成戴谈宋茅庞熊纪舒屈项祝董梁"
              "杜阮蓝闵席季麻强贾路娄危江童颜郭梅盛林刁钟徐邱骆高夏蔡田樊胡凌霍"
              "万柯卢莫房裘缪干解应宗丁宣贲邓郁单杭洪包诸左石崔吉钮龚程嵇邢滑裴"
              "陆荣翁荀羊於惠甄曲家封芮羿储靳汲邴糜松井段富巫乌焦巴弓牧隗山谷车"
              "侯宓蓬全郗班仰秋仲伊宫宁仇栾暴甘钭厉戎祖武符刘景詹束龙叶幸司韶"
              "郜黎蓟薄印宿白怀蒲台从鄂索咸籍赖卓蔺屠蒙池乔阴郁胥能苍双闻莘党"
              "翟谭贡劳逄姬申扶堵冉宰郦雍却璩桑桂濮牛寿通边扈燕冀郏浦尚农温别"
              "庄晏柴瞿阎充慕连茹习宦艾鱼容向古易慎戈廖庚终暨居衡步都耿满弘匡"
              "国文寇广禄阙东殴殳沃利蔚越夔隆师巩厍聂晁勾敖融冷訾辛阚那简饶空"
              "曾毋沙乜养鞠须丰巢关蒯相查后荆红游竺权逯盖益桓公")

# 姓名识别：姓氏 + 1~2 个汉字，依赖语境以抑制纯中文文本极高误报率。
# 分两级语境：
#  - 强语境（标签词 姓名/联系人/客户/借款人/担保人… 或 冒号 ：）：前缀已提供强上下文，
#    放宽后缀边界（仅按 姓氏+1~2 字 定长），提升召回（如"客户张三于..."可命中"张三"）。
#  - 弱语境（逗号/顿号/空格/括号/句号等标点前缀）：保留后缀边界（称谓或标点/结尾），
#    防止自由文本误报。前后边界同时覆盖 中英文标点（全角/半角），避免混排漏检。
# 中英文标点边界集（全角/半角一并包含）：空白、逗号、句号、冒号、分号、叹号、问号、顿号、
# 括号（全/半角）。供弱语境的前顾与后顾复用。
_CN_BOUND = r"\s,，.:：;；!！?？、()（）"
_CN_NAME_STRONG = re.compile(
    r"(?:(?<=姓名)|(?<=联系人)|(?<=客户)|(?<=经办人)|(?<=申请人)|"
    r"(?<=被谈话人)|(?<=被询问人)|(?<=借款人)|(?<=担保人)|"
    r"(?<=[:：]))"
    r"(?P<name>(?:%s|[%s])[一-龥]{1,2})"
    % ("|".join(_CN_COMPOUND), _CN_SINGLE)
)
_CN_NAME_WEAK = re.compile(
    r"(?<=[%s])"
    r"(?P<name>(?:%s|[%s])[一-龥]{1,2})"
    r"(?=(?:先生|女士|同学|老师|经理|总监|主任|教授|工|君|[%s]|$))"
    % (_CN_BOUND, "|".join(_CN_COMPOUND), _CN_SINGLE, _CN_BOUND)
)
CN_NAME_PATTERNS = [_CN_NAME_STRONG, _CN_NAME_WEAK]

# 地址识别：对齐 summarize 三强信号分支（pii_precheck commit f3b53cf），消除单字触发词误报。
# 旧版把「路/市/区/号/中心/道」等高频字**单独**作触发词，导致「电路/道路/中心思想/数据中心」等
# 普通叙述被整段误判为地址（开启 --cn-enhance 即复现，详见「地址误报反馈_DESEN排查.md」）。
# 现改为复合强信号，四类覆盖：
#  ① 路/街/巷/道/大道/弄 + 门牌数字 → 覆盖「中关村大街1号」「人民路100号」（无市/区前缀）；
#  ② 省/自治区 + 下级行政区（市/区/县/旗/盟/镇/街道/乡）→ 覆盖「广东省深圳市南山区」；
#  ③ 市 + 区/县/旗/盟 → 覆盖「北京市朝阳区」「杭州市西湖区」（无楼栋）；
#  ④ 特别行政区（独立强信号）→ 覆盖「香港特别行政区」「澳门特别行政区」。
# 分支②③后可附带「路/楼栋/小区」等余下地址要素（_CN_ADDR_TAIL），一次性吞掉整段地址
# （如「广东省深圳市龙岗区坂田街道XX小区3栋502室」「北京市朝阳区建国路88号」），保证单匹配、整段脱敏；
# 中英文标点天然作为截断边界；门牌分支强制「路/街后跟数字」，从源头消除「技术路线/道路」类误报。
# ⚠️ 尾缀仅在有效基匹配之后延伸，不会制造新误报；🔴 高频普通字（中心/道单字/村/乡/镇）已排除出独立触发。
_CN_ADDR_TAIL = (
    r"(?:"
    r"[一-龥0-9A-Za-z\-]{1,6}"
    r"(?:省|市|自治区|区|县|旗|盟|路|街道|街|道|大道|巷|弄|"
    r"号|栋|幢|单元|室|小区|花园|广场|大厦|公寓|村|镇|乡)"
    r")*"
)
_CN_ADDRESS = re.compile(
    # ① 路+门牌（无市/区前缀，如 中关村大街1号 / 人民路100号）
    r"[一-龥]{2,10}(?:路|街|巷|道|大道|弄)\s?\d{1,6}(?:号|栋|幢|单元|室)?"
    + _CN_ADDR_TAIL
    + r"|"
    # ② 省/自治区 + 下级行政区（可附带路/楼栋详情，如 广东省深圳市龙岗区坂田街道XX小区3栋502室）
    r"[一-龥]{2,6}(?:省|自治区)[一-龥]{2,8}(?:市|区|县|旗|盟|镇|街道|乡)"
    + _CN_ADDR_TAIL
    + r"|"
    # ③ 市 + 区/县（可附带路/楼栋详情，如 北京市朝阳区建国路88号）
    r"[一-龥]{2,6}市[一-龥]{2,6}(?:区|县|旗|盟)"
    + _CN_ADDR_TAIL
    + r"|"
    # ④ 特别行政区（独立强信号，如 香港特别行政区）
    r"[一-龥]{2,6}特别行政区"
)

# 机构识别：含 公司/银行/医院/大学 等组织后缀
_CN_ORG = re.compile(
    r"[一-龥]{2,}(?:公司|集团|银行|医院|大学|学院|学校|研究所|研究院|局|部|"
    r"委员会|协会|基金会|事务所|工厂|超市|店|分行|支行|证券|保险|基金|信托|"
    r"律所|合作社|中心|平台)"
)

CN_ENHANCE_PATTERNS = {
    "cn_name": CN_NAME_PATTERNS,
    "cn_address": _CN_ADDRESS,
    "cn_org": _CN_ORG,
}


def build_patterns(cn_enhance: bool):
    """返回生效的正则集合（cn_enhance 时并入中文增强）。"""
    pats = dict(PATTERNS)
    if cn_enhance:
        pats.update(CN_ENHANCE_PATTERNS)
    return pats


# 代码场景：密钥 / Token 类（仅对代码与配置文件生效）
# 值支持三种形态：单引号 '...'、双引号 "..."、或无引号（遇空白/引号截止）。
# 注意：引号型必须同时匹配「开头 + 结尾」引号，否则替换时结尾引号会残留（如 password='***'）。
SECRET_PATTERN = re.compile(
    r"""(?i)(['"]?)(?:api[_-]?key|secret|token|password|passwd|access[_-]?key|private[_-]?key|auth)['"]?\s*[:=]\s*(?P<val>'[^']{4,}'|"[^"]{4,}"|[^\s'"]{4,})"""
)

# SQL INSERT 位置参数口令启发式：数据库 dump 中口令常以
#   INSERT INTO users (..., password, ...) VALUES (..., 'secret_pass_123', ...)
# 位置参数形式出现，值前无 password= 前缀，SECRET_PATTERN 无法捕获。
# 此处按"列名含口令关键词 → 对应位置值"定位并脱敏（列名缺失/解析失败时静默跳过）。
_SQL_INS_RE = re.compile(
    r"INSERT\s+INTO\s+[^\s(;]+(?:\s*\(([^)]*)\))?\s*VALUES\s*\(([^)]*)\)",
    re.I | re.S)
_SQL_SECRET_COL = re.compile(
    r"(?:passwd|password|pwd|secret|token|api[_-]?key|access[_-]?key|private[_-]?key|auth)",
    re.I)


def _sql_insert_secret_values(text):
    """返回 [(带引号原值, 纯净值)]：INSERT ... (含口令列) VALUES 中对应位置的值。

    用 csv 解析 VALUES（正确处理引号内逗号）；列名缺失（INSERT INTO t VALUES ...）
    时无法定位口令列，跳过（保守，避免误伤）。"""
    import csv
    import io
    out = []
    for m in _SQL_INS_RE.finditer(text):
        cols_str, vals_str = m.group(1), m.group(2)
        if not cols_str:
            continue
        cols = [c.strip().strip('`"[]') for c in cols_str.split(",")]
        try:
            vals = next(csv.reader(io.StringIO(vals_str)))
        except Exception:
            continue
        for i, col in enumerate(cols):
            if i >= len(vals):
                break
            if _SQL_SECRET_COL.search(col):
                v = vals[i].strip()
                if len(v) >= 4:
                    inner = (v[1:-1] if len(v) >= 2 and v[0] in "'\"" and v[-1] == v[0]
                             else v)
                    if inner:
                        out.append((v, inner))
    return out

CODE_EXTS = {".py", ".js", ".ts", ".go", ".java", ".c", ".cpp", ".rb", ".php",
              ".yml", ".yaml", ".toml", ".env", ".ini", ".sh", ".sql"}
# 密钥检测扩展：HTML 页面内嵌的 <script> 中常见 adminToken/secret 等赋值，
# 是真实高频泄漏点，故 html/htm 一并启用 SECRET_PATTERN 密钥检测（分类仍属 TEXT）。
# 跨境电商高频把 API key/access_token 写在 CSV 导出、JSON 配置里（GA token、Stripe
# api_key 等），故 csv/json 一并启用 SECRET_PATTERN（如 csv 的 access_token=…、json 的
# "api_key":"…" 形态），堵住数据文件内密钥明文留存。
SECRET_EXTS = CODE_EXTS | {".html", ".htm", ".csv", ".json"}
TEXT_EXTS = {".txt", ".md", ".csv", ".tsv", ".json", ".log", ".xml", ".html",
             ".htm", ".rst", ".eml"}
# 二进制/库文件：抽取文本后再脱敏（抽取库缺失时跳过并告警，不崩溃）
EXTRACT_EXTS = {".docx", ".xlsx", ".pptx", ".pdf", ".db"}
# 影像/图片：脚本不做 OCR，必须提醒用户本地 OCR 转文本后再纳入
IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".tif", ".tiff",
              ".webp", ".heic", ".heif"}
# skip manifest 的明确可执行提醒（让用户知道“为什么没处理、下一步做什么”）
REMIND = {
    "encrypted": "加密文档：脚本无法处理加密文件，请先运行 preprocess 自动解密"
                 "（提供 --passwords-file），或本地手动解密后再纳入脱敏",
    "image_only": "纯图片型/扫描件 PDF：本文件无文本层，请先运行 preprocess"
                  "（内置 rapidocr 本地 OCR，完全离线）识别为文本后再纳入脱敏",
    "empty": "文档未提取到文本（可能为图片型扫描件或空文档）：若含敏感信息，"
             "请先运行 preprocess 做 OCR 后再纳入脱敏",
    "error": "文档读取/抽取失败（可能已加密或文件损坏）：请先解密或确认文件完整性后再纳入脱敏",
    "no_lib": "缺少抽取库，无法处理该二进制文档：请安装对应库（docx→python-docx，"
              "xlsx→openpyxl，pptx→python-pptx，pdf→pypdfium2）后再处理",
}

# “不得外传”红线警告（按类别显式给出，杜绝后续流程把原文件/未脱敏副本送出去）
# 说明：以下三类严禁离本地，是本地预处理关卡的核心约束，会被 preprocess / run / audit 反复提示。
WARN_ENCRYPTED = ("⚠️ 严禁外传：解密密码、加密原文件、以及“已解密但未脱敏”的副本——"
                  "三者均不得发送至任何外部/云端。仅在本地完成解密，再用本工具脱敏后，"
                  "方可上云脱敏副本。")
WARN_IMAGE = ("⚠️ 严禁外传：原始图片/图片型PDF，以及 OCR 识别出、未经脱敏的文本——"
              "二者均不得发送至任何外部/云端。仅在本地完成 OCR 与脱敏后，"
              "方可上云脱敏副本。")
WARN_GENERIC = ("⚠️ 该文件未经处理，上云前须先本地转为文本/OCR/解密或确认不含敏感信息；"
                "原始文件与未脱敏文件均不得直接外传。")

# ---------------------------------------------------------------------------
# 数据清洗建议（提醒"先清洗再脱敏"）
# 小微企业的数据填写/传递往往随意：手机号带空格/横线、少一位、15 位旧身份证、
# 订单号/流水号与银行卡区间重叠等。这些形态要么漏检、要么误报，
# 脚本不强行处理（避免误伤），而是检测并**明确提醒**用户先做字段级清洗/确认。
# 清洗建议只提示、不脱敏（fail-safe：宁可提醒也不静默放过）。
# ---------------------------------------------------------------------------
MESSY_PATTERNS = {
    # 带分隔符的手机号（138-0013-8004 / 138 0013 8002 / +86 138 0013 8003）
    "messy_phone_sep": re.compile(
        r"(?<![0-9A-Za-z])(?:\+?86[-\s]?)?1[3-9]\d[-\s]\d{3,4}[-\s]\d{3,4}(?!\d)"),
    # 10 位手机号（少一位，录错/占位常见）
    "messy_phone_short": re.compile(r"(?<![0-9A-Za-z])1[3-9]\d{8}(?![0-9A-Za-z])"),
    # 15 位旧身份证（1999 年前签发，老员工/旧单存量）
    "messy_id15": re.compile(
        r"(?<![0-9A-Za-z])[1-9]\d{5}\d{2}(?:0[1-9]|1[0-2])(?:0[1-9]|[12]\d|3[01])\d{3}(?![0-9A-Za-z])"),
    # 订单号/流水号/计划ID 标签后紧跟 16-19 位纯数字（与银行卡区间重叠，易误判）
    "messy_order_like": re.compile(
        r"(?:订单号|订单编号|单号|流水号|计划\s*ID|编号|单据号)[:：=\s]{0,4}\d{16,19}"),
}
CLEANING_ADVICE_TEXT = (
    "⚠️ 清洗建议：检测到疑似「未清洗/不规范」数据形态（带分隔符或短位手机号、"
    "15 位旧身份证、订单号/流水号与银行卡区间重叠等）。上述形态未被自动脱敏或可能被误判，"
    "请先做字段级数据清洗（去空格/横线、补齐位数、区分订单号与银行卡）后再脱敏，"
    "或人工复核确认。切勿在未清洗状态下直接上云。")


def _find_cleaning_advice(text):
    """检测疑似未清洗数据形态，返回 {类别: 出现次数}（不脱敏，仅提醒）。"""
    return {k: len(p.findall(text)) for k, p in MESSY_PATTERNS.items()
            if p.findall(text)}


# 预处理/脱敏流程自身产出的元数据文件（不含业务敏感信息，run 时须跳过、不可当作业务文本脱敏）
META_SKIP = {
    "desensitize_preprocess.json",
    "desensitize_preprocess_summary.md",
    "desensitize_report.json",
    "desensitize_audit.md",
    # 工作区（统一成果中心）自带的索引/说明文件，非业务数据，须跳过
    "成果索引.json",
    "成果索引.md",
    "00_使用说明.md",
    "05_上云前自检报告.md",
    # 以下为“公开声明”机制的伴随清单控制文件，本身非业务数据，须跳过
    ".nodesens",
    "desensitize_manifest.json",
}
